from openpyxl import Workbook
import openpyxl

def importExcel(workbook):
    book = openpyxl.load_workbook(workbook)
    sheet = book.active

    # Find the number of rows in the excel sheet
    rowCount = sheet.max_row+1
    columnCount = sheet.max_column+1

    # A 2D array, containing a list of the contents of each row
    sheet_contents = []

    # Iterates through each row (Except for the head
    # Changes the date strings into date format

    for r in range(1,rowCount):
        row_contents = []
        for c in range(1,columnCount):
            row_contents.append(sheet.cell(row=r,column=c).value)
        sheet_contents.append(row_contents)

    return sheet_contents